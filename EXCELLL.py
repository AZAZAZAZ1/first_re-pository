
#C:\Users\E.H.PAUE\Desktop\python lessons 

import openpyxl

import os

os.chdir('C:\\Users\\E.H.PAUE\\python lessons')# change the directory of where the file is Excell sheet saved  # dont forget to put \\ in the path


workbook = openpyxl.load_workbook('example.xlsx') # to open the excell sheet that saved in the above directory  # the type of (woorkbook) is 'worbook'
type(workbook)

sheet= workbook["Sheet1"] # to open specific sheet
type(sheet)

sheetnames= workbook.sheetnames # to get all sheets in the workbook

#oldcode   :sheet = workbook.get_sheet_by_name('Sheet1') # to open specific sheet in the excell work book   # type is 'worksheet'
# old code :sheet = workbook.get_sheet_names()  # to get all shhets in the workbook
# old code :sheet['A1'] # cell A1
#Old code  :cell.value() # to get the cell value / or what written on it.

print (type(workbook))
print(type(sheet))

cell=sheet['A1'] # to select specif cell only 
print(cell)# you will get <Cell 'Sheet1'.A1>

cell_value= cell.value # to get the value of the cell ('A1')
print(cell_value)  # you will get 73

cellobject = sheet.cell(2,6) # row 2, column 6   you will get F2 (te select the cell by knowing the row and colum numbers)
print(cellobject)

for i in range(1,8):
    print (i,sheet.cell(row=1, column=2).value)  #


    
    
#print(sheet)
#print(sheetnames)

#cell_value = sheet.cell(1,5) # column1 , row 5
#sheet.value
#print(cell_value) 
#print (sheet.value)
#wb.close()

# another code that can be used :
#from openpyxl import *
#source_file = load_workbook("file_path") # you can use the path instead of file name, no need to change the directory

#sheet = source_file["sheet_name"]
