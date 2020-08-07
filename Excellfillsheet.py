
#C:\Users\E.H.PAUE\Desktop\python lessons 

import openpyxl
from openpyxl import Workbook # becairful capital W for Workbook
import os

import xlsxwriter
workbook = xlsxwriter.Workbook('C:\\Users\\E.H.PAUE\python lessons\\Welocme.xlsx')# create workbook(file) in this directory called: Welcome
worksheet = workbook.add_worksheet('sheta')# creat sheet called "sheta"
worksheet.write('A1', 'Welcome')# write welcome in cell A1
worksheet.write('B1', 'To')     # write To in cell B1
worksheet.write('C1', 'Python') # write python in C1

workbook.close()


#reference: https://www.sqlshack.com/python-scripts-to-format-data-in-microsoft-excel/


#os.chdir('C:\\Users\\E.H.PAUE\\python lessons')# change the directory of where the file is Excell sheet saved  # dont forget to put \\ in the path

#wb = openpyxl.Workbook() # to create new excel file # becairful capital W for Workbook

#sheet = Workbook.active

#sheetnames= wb.sheetnames
#print(sheetnames)

#sheet ['A1']=43
#sheet["A1"].value == None
#sheet["B1"].value = 
#Workbook.save(filename="hello_world.xlsx")

#sheet= wb["Sheet1"] # to open specific sheet


#sheetnames= wb.sheetnames # to get all sheets in the workbook


